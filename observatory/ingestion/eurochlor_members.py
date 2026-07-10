"""Production & capacity agent, member leg — Euro Chlor survey workbooks (Phase 2).

Reads the monthly statistics report and annual survey report dropped into
data/eurochlor_drop/. AUTHORITATIVE for EU capacity/production/utilisation
(config/source_of_truth.yaml); the public eurochlor_web scrape remains the
freshness cross-check.

Every row is tagged redistribution_class='licensed' — member-restricted,
served only while the member gate (dashboard password) is active. Blank/na/**
cells are competition-rule suppressions: skipped and counted, never zero.
raw_landing receives the file manifest only, not the licensed content.
"""
import re
from datetime import datetime, timezone

import openpyxl

from observatory.ingestion.base import IngestionAgent
from observatory.ingestion.periods import period_start
from observatory.provenance import SeriesRow
from observatory.settings import DATA_DIR

DROP_DIR = DATA_DIR / "eurochlor_drop"
GEO = "EU27_EFTA_UK"   # 'ALL Euro Chlor countries'

STOCK_BANDS = {8: "TOTAL", 9: "MERCURY_MEMBRANE", 10: "DIAPHRAGM",
               11: "LIQUIDS", 12: "SOLIDS"}   # 0-based col -> band

CL2_SECTIONS = {   # sheet 'Cl2' section prefix -> (series, unit, scale)
    "A. TOTAL NAME PLATE CAPACITY": ("production.capacity", "t Cl2/yr", 1000.0),
    "B. PRODUCTION AT END OF PERIOD": ("production.production", "t Cl2", 1000.0),
    "C. UTILISATION RATE": ("production.utilisation", "%", 100.0),
    "D. APPARENT NATIONAL CONSUMPTION": ("consumption.cl2_apparent", "t Cl2", 1000.0),
}


def _num(v):
    """Numeric value or None for competition-rule suppressions (blank/na/**)."""
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _year(v):
    """Parse a year from a header cell like 2016, '2015 \\n5 groups'."""
    s = str(v).strip()
    m = re.match(r"(20\d\d)", s)
    return int(m.group(1)) if m else None


def _group_id(label: str) -> str:
    return "ECG_" + re.sub(r"[^A-Z0-9]+", "_", label.upper()).strip("_")[:40]


class EuroChlorMemberAgent(IngestionAgent):
    name = "eurochlor_members"
    source = "Euro Chlor member surveys (restricted)"

    def available(self):
        if not list(DROP_DIR.rglob("*.xlsx")):
            return False, f"no survey workbooks under {DROP_DIR}"
        return True, ""

    def fetch(self):
        files = sorted(DROP_DIR.rglob("*.xlsx"))
        def kind(name):
            low = name.lower()
            if "monthly" in low:
                return "monthly"
            if "annual survey" in low:
                return "annual"
            if "chlorine production" in low:
                return "regions"   # US (Chlorine Institute) + China (CCAIA) utilisation
            return "unknown"
        return [(str(f), {"file": str(f), "kind": kind(f.name)}) for f in files]

    def parse(self, payloads):
        retrieved_at = datetime.now(timezone.utc)
        self._groups = {}
        self._suppressed = 0
        rows = []
        for _, meta in payloads:
            if meta["kind"] == "monthly":
                rows += self._parse_monthly(meta["file"], retrieved_at)
            elif meta["kind"] == "annual":
                rows += self._parse_annual(meta["file"], retrieved_at)
            elif meta["kind"] == "regions":
                rows += self._parse_regions(meta["file"], retrieved_at)
        return rows

    def _row(self, series, period, value, unit, src, retrieved_at, band=None, geo=GEO):
        return SeriesRow(
            series_id=series, geo_id=geo, period=str(period),
            period_start=period_start(str(period)), value=value, unit=unit,
            band=band, source="Euro Chlor member survey", source_dataset=src,
            reference_period=str(period), retrieved_at=retrieved_at,
            redistribution_class="licensed")

    # ---------- monthly report ----------
    def _parse_monthly(self, path, retrieved_at):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Data used for graphs"]
        src = f"Monthly Statistics Report [{path.split('/')[-1]}]"
        rows = []
        for r in ws.iter_rows(min_row=4, values_only=True):
            # chlorine block: c15 date, c17 monthly production t (daily x actual days)
            if isinstance(r[14], datetime) and _num(r[16]) is not None:
                rows.append(self._row("production.production", r[14].strftime("%Y-%m"),
                                      _num(r[16]), "t Cl2", src, retrieved_at))
            # caustic production block: c19 date, c21 monthly production t
            if isinstance(r[18], datetime) and _num(r[20]) is not None:
                rows.append(self._row("production.caustic_production", r[18].strftime("%Y-%m"),
                                      _num(r[20]), "t NaOH", src, retrieved_at))
            # caustic stocks: c8 date, c9-13 kt by technology/form
            if isinstance(r[7], datetime):
                for col, band in STOCK_BANDS.items():
                    v = _num(r[col])
                    if v is not None:
                        rows.append(self._row("production.caustic_stocks", r[7].strftime("%Y-%m"),
                                              v * 1000, "t NaOH", src, retrieved_at, band=band))
        wb.close()
        return rows

    # ---------- annual survey ----------
    def _parse_annual(self, path, retrieved_at):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        src = f"Annual Survey Final Report [{path.split('/')[-1]}]"
        rows = []
        rows += self._parse_cl2(wb, src, retrieved_at)
        rows += self._parse_cl2_type(wb, src, retrieved_at)
        rows += self._parse_cl2_uses(wb, src, retrieved_at)
        rows += self._parse_naoh_geo(wb, src, retrieved_at)
        wb.close()
        return rows

    def _parse_cl2(self, wb, src, retrieved_at):
        """Sections A-D: country-group rows x year columns. Levels part only
        (indices restatement starts ~row 106)."""
        grid = list(wb["Cl2"].iter_rows(min_row=1, max_row=101, values_only=True))
        rows, years, section = [], None, None
        for r in grid:
            label = str(r[0]).strip() if r[0] is not None else ""
            hit = next((v for k, v in CL2_SECTIONS.items() if label.startswith(k)), None)
            if hit:
                section = hit
                continue
            if years is None and r[1] is not None and _year(r[1]) == 2000:
                years = [_year(c) for c in r[1:]]
                continue
            if not (section and years and label):
                continue
            if label.startswith(("Realised", "BLANK", "(*)", "Dark", "Chlorine", "Indices")):
                continue
            series, unit, scale = section
            geo = GEO if label == "TOTAL" else _group_id(label)
            if label != "TOTAL":
                self._groups[geo] = f"Euro Chlor group: {label}"
            for year, cell in zip(years, r[1:]):
                if year is None:
                    continue
                v = _num(cell)
                if v is None:
                    self._suppressed += 1
                    continue
                rows.append(self._row(series, year, v * scale, unit, src, retrieved_at, geo=geo))
        return rows

    def _parse_cl2_type(self, wb, src, retrieved_at):
        """Technology shares (fractions) x years."""
        grid = list(wb["Cl2 type"].iter_rows(min_row=1, max_row=32, values_only=True))
        rows, years = [], None
        techs = ("Diaphragm", "Mercury", "Membrane", "Fused", "Other")
        for r in grid:
            label = str(r[0]).strip().split("\n")[0] if r[0] is not None else ""
            if years is None and r[1] is not None and _year(r[1]):
                years = [_year(c) for c in r[1:]]
                continue
            if not years or not label.startswith(techs):
                continue
            for year, cell in zip(years, r[1:]):
                if year is None:
                    continue
                v = _num(cell)
                if v is None:
                    continue   # tech absent that year — not a suppression
                rows.append(self._row("production.tech_share", year, v * 100, "%",
                                      src, retrieved_at, band=label.split()[0].upper()))
        return rows

    def _parse_cl2_uses(self, wb, src, retrieved_at):
        """Applications x years; volume columns are those whose r5 header is a year
        (Share/Indices columns in between are skipped)."""
        grid = list(wb["Cl2 Uses "].iter_rows(min_row=5, max_row=32, values_only=True))
        header = grid[0]
        year_cols = [(j, _year(c)) for j, c in enumerate(header) if _year(c)]
        rows = []
        for r in grid[1:]:
            label = (str(r[1]).strip() if r[1] is not None else
                     str(r[0]).strip() if r[0] is not None else "")
            label = label.split("\n")[0]
            if not label or label.startswith(("BLANK", "NA:", "Chorin", "Chlorin")):
                continue
            band = "TOTAL" if label.lower().startswith("total") else label[:60]
            for j, year in year_cols:
                v = _num(r[j])
                if v is None:
                    self._suppressed += 1
                    continue
                rows.append(self._row("consumption.cl2_by_use", year, v, "t Cl2",
                                      src, retrieved_at, band=band))
        return rows

    def _parse_naoh_geo(self, wb, src, retrieved_at):
        """Sales / captive-use sections: group rows x years. Two 2015 columns
        exist ('3 groups' and '5 groups' basis) — the later one wins."""
        ws = wb["NaOH Geo (2007-2025)"]
        grid = list(ws.iter_rows(min_row=1, values_only=True))
        rows, years, band = [], None, None
        for r in grid:
            if years is None:
                cand = [_year(c) for c in r[2:]]
                if sum(1 for y in cand if y) >= 5:
                    years = cand
                continue
            c1 = str(r[0]).strip() if r[0] is not None else ""
            label = str(r[1]).strip().split("\n")[0] if r[1] is not None else ""
            if re.match(r"^\d+\.$", c1):   # section header row (carries the totals)
                low = label.lower()
                band = ("SALES" if "sales" in low else
                        "CAPTIVE" if "captive" in low else low[:30].upper() or "OTHER")
                geo, is_group = GEO, False
            elif label and years:
                if "should equal" in label.lower() or label.lower().startswith("total sodium"):
                    continue
                geo, is_group = _group_id(label), True
            else:
                continue
            if is_group:
                self._groups[geo] = f"Euro Chlor group: {label}"
            seen_years = {}
            for idx, (year, cell) in enumerate(zip(years, r[2:])):
                if year is None:
                    continue
                seen_years[year] = cell   # later duplicate (5-groups basis) wins
            for year, cell in seen_years.items():
                v = _num(cell)
                if v is None:
                    self._suppressed += 1
                    continue
                rows.append(self._row("consumption.naoh_apparent", year, v, "t NaOH",
                                      src, retrieved_at, band=band, geo=geo))
        return rows

    def _parse_regions(self, path, retrieved_at):
        """'compared with other regions': monthly utilisation for Euro Chlor
        (extends the EU series back to 2000) and the US (Chlorine Institute);
        China (CCAIA) carries one annual value repeated across the year's
        monthly rows — stored as annual."""
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["compared with other regions"]
        fname = path.split("/")[-1]
        rows = []
        cn_by_year = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not isinstance(r[0], datetime):
                continue
            period = r[0].strftime("%Y-%m")
            eu, us, cn = _num(r[1]), _num(r[2]), _num(r[3])
            if eu is not None:
                rows.append(SeriesRow(
                    series_id="production.utilisation", geo_id=GEO, period=period,
                    period_start=period_start(period), value=eu * 100, unit="%",
                    source="Euro Chlor member survey",
                    source_dataset=f"compared with other regions [{fname}]",
                    reference_period=period, retrieved_at=retrieved_at,
                    redistribution_class="licensed"))
            if us is not None:
                rows.append(SeriesRow(
                    series_id="production.utilisation", geo_id="US", period=period,
                    period_start=period_start(period), value=us * 100, unit="%",
                    source="Chlorine Institute (via Euro Chlor)",
                    source_dataset=f"compared with other regions [{fname}]",
                    reference_period=period, retrieved_at=retrieved_at,
                    redistribution_class="licensed"))
            if cn is not None:
                cn_by_year.setdefault(r[0].strftime("%Y"), []).append(cn)
        for year, vals in sorted(cn_by_year.items()):
            # 'only yearly figures': the annual value is repeated monthly
            rows.append(SeriesRow(
                series_id="production.utilisation", geo_id="CN", period=year,
                period_start=period_start(year), value=sum(vals) / len(vals) * 100,
                unit="%", source="CCAIA (via Euro Chlor)",
                source_dataset=f"compared with other regions [{fname}] — annual figure",
                reference_period=year, retrieved_at=retrieved_at,
                redistribution_class="licensed"))
        wb.close()
        return rows

    def pre_insert(self, conn, rows):
        for geo_id, name in self._groups.items():
            conn.execute(
                """INSERT INTO dim_geo (geo_id, name, kind) VALUES (%s, %s, 'region')
                   ON CONFLICT (geo_id) DO NOTHING""", (geo_id, name))

    def run(self):
        result = super().run()
        suppressed = getattr(self, "_suppressed", 0)
        if suppressed and isinstance(result, dict):
            result["suppressed_cells"] = suppressed
        return result
